"""
RAPPORT NUTRITION - Script Python exécutable avec python {MODULE}.py
"""

import pandas as pd
import numpy as np
import os
import re
import time
import warnings
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse
from dotenv import load_dotenv
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import xlsxwriter
import pymysql
from sqlalchemy import create_engine
from selenium import webdriver


print("="*60)
today_date = pd.to_datetime('today')
print(f"DEBUT DE LA PIPELINE DE NUTRITION à {today_date}")
print("="*60)


#=================================================================================================
def extraire_data(df, start_date, end_date, date_col):
    # Conversion des dates d'entrée
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)

    # Conversion de la colonne de date dans le DataFrame
    if date_col in df.columns:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    else:
        raise ValueError(f"Colonne '{date_col}' absente du DataFrame.")

    # Filtrage
    df = df[df[date_col].between(start_date, end_date)]

    return df
#=============================================
import pandas as pd

def clean_column_names(df, expr_to_remove):
    """
    Supprime une expression donnée au début des noms de colonnes d'un DataFrame.

    Args:
        df (pd.DataFrame): le dataframe cible
        expr_to_remove (str): l'expression à supprimer (ex: 'form.depistage.')

    Returns:
        pd.DataFrame: un DataFrame avec colonnes renommées
    """
    df = df.rename(columns=lambda col: col.replace(expr_to_remove, ""))
    return df
#========================================================================================
import pandas as pd

def create_binary_symptom_columns(df, col="symptoms"):
    """
    Transforme une colonne de type string en colonnes indicatrices binaires (0/1).
    - Si 'aucun' est présent avec d'autres modalités => on ignore 'aucun'.
    - Si 'aucun' est la seule modalité => on crée une colonne 'aucun'=1.
    
    Args:
        df (pd.DataFrame): le DataFrame source
        col (str): nom de la colonne contenant les modalités séparées par espaces

    Returns:
        pd.DataFrame: DataFrame avec colonnes indicatrices ajoutées
    """
    # Créer une liste de sets de symptômes par ligne
    df[col] = df[col].fillna("").str.lower()
    df[col] = df[col].str.strip()
    df[col] = df[col].str.split()

    # Extraire toutes les modalités distinctes
    all_modalities = set(mod for mods in df[col] for mod in mods)

    # On retire 'aucun' provisoirement
    if "aucun" in all_modalities:
        all_modalities.remove("aucun")

    # Créer colonnes indicatrices pour chaque modalité (hors 'aucun')
    for mod in all_modalities:
        df[mod] = df[col].apply(lambda mods: 1 if mod in mods else 0)

    # Gérer la règle spéciale pour 'aucun'
    df["aucun"] = df[col].apply(lambda mods: 1 if mods == ["aucun"] else 0)

    return df
#============================================================================
def get_age_in_year(df, dob_col):
    """
    Calcule l'âge en années pour chaque ligne d'un DataFrame où la colonne date de naissance n'est pas NaN.
    
    Args:
        df (pd.DataFrame): le DataFrame d'entrée
        dob_col (str): nom de la colonne contenant la date de naissance
        
    Returns:
        pd.DataFrame: DataFrame avec une nouvelle colonne 'age_years'
    """
    # Faire une copie du DataFrame pour éviter les modifications sur l'original
    df_copy = df.copy()
    
    # Vérifier que la colonne existe
    if dob_col not in df_copy.columns:
        raise ValueError(f"Colonne '{dob_col}' absente du DataFrame.")
    
    # Convertir la colonne date de naissance en datetime
    df_copy[dob_col] = pd.to_datetime(df_copy[dob_col], errors='coerce')
    
    # Calculer l'âge seulement là où dob_col n'est pas NaN
    today = pd.Timestamp.now()
    
    # Créer la colonne age_years avec des NaN par défaut
    df_copy['age_years'] = pd.NaT
    
    # Calculer l'âge seulement pour les lignes où date de naissance n'est pas NaN
    mask = df_copy[dob_col].notna()
    df_copy.loc[mask, 'age_years'] = (
        (today - df_copy.loc[mask, dob_col]).dt.days / 365.25
    ).round(0)
    df_copy['age_years'] = df_copy['age_years'].fillna(0).astype(int)
    return df_copy

#=====================================================================
def get_age_in_months(df, dob_col):
    """
    Calcule l'âge en mois pour chaque ligne d'un DataFrame où la colonne date de naissance n'est pas NaN.
    
    Args:
        df (pd.DataFrame): le DataFrame d'entrée
        dob_col (str): nom de la colonne contenant la date de naissance
        
    Returns:
        pd.DataFrame: DataFrame avec une nouvelle colonne 'age_months'
    """
    # Faire une copie du DataFrame pour éviter les modifications sur l'original
    df_copy = df.copy()
    
    # Vérifier que la colonne existe
    if dob_col not in df_copy.columns:
        raise ValueError(f"Colonne '{dob_col}' absente du DataFrame.")
    
    # Convertir la colonne date de naissance en datetime
    df_copy[dob_col] = pd.to_datetime(df_copy[dob_col], errors='coerce')
    
    # Calculer l'âge seulement là où dob_col n'est pas NaN
    today = pd.Timestamp.now()
    
    # Créer la colonne age_months avec des NaN par défaut
    df_copy['age_months'] = pd.NaT
    
    # Calculer l'âge en mois seulement pour les lignes où date de naissance n'est pas NaN
    mask = df_copy[dob_col].notna()
    df_copy.loc[mask, 'age_months'] = (
        (today - df_copy.loc[mask, dob_col]).dt.days / 30.44  # 30.44 = moyenne de jours par mois
    ).round(0)
    
    # Convertir en entier, remplacer NaN par 0
    df_copy['age_months'] = df_copy['age_months'].fillna(0).astype(int)
    
    return df_copy
#=====================================================================
import pandas as pd

def create_normalized_health_index(df):
    """
    Crée un indice pondéré et normalisé (0-10) basé sur symptômes et variables nutritionnelles,
    puis ajoute une catégorisation de risque.
    """

    # Pondération malnutrition
    df["malnutrition_score"] = df["manutrition_type"].map(
        {"MAM": 2, "MAS": 4}
    ).fillna(0)

    # Pondération diarrhée
    df["diarrhea_score"] = df["diarrhea"].map(
        {"yes": 3, "no": 0}
    ).fillna(0)

    # Pondération oedème
    df["oedem_score"] = df["edema"].map(
        {"yes": 5, "no": 0}
    ).fillna(0)
    
        # Pondération oedème
    df["lesion_cutane_score"] = df["lesion_cutane"].map(
        {"yes": 3, "no": 0}
    ).fillna(0)

    # Score brut pondéré
    df["health_index_raw"] = (
        df["toux"] * 1 +
        df["fievre"] * 2 +
        df["douleurs_abdominales"] * 2 +
        df["vomissements"] * 2 +
        df["malnutrition_score"] +
        df["diarrhea_score"] +
        df["oedem_score"]+
        df["lesion_cutane_score"]
    )

    # Maximum théorique possible
    max_score = 22
    df["health_index_norm"] = (df["health_index_raw"] / max_score) * 10

    # Catégorisation
    def categorize(score):
        if score <= 3.0:
            return "Faible risque"
        elif score <= 6.0:
            return "Risque modéré"
        else:
            return "Haut risque"

    df["risk_category"] = df["health_index_norm"].apply(categorize)

    return df
# df[["toux","fievre","vomissement","manutrition_type","edema","lesion_cutane","diarrhea","douleurs_abdominales","health_index_raw","health_index_norm","risk_category"]]
#====================================================================
import pandas as pd
from typing import Optional

def filter_patients(
    df: pd.DataFrame,
    date_threshold: str = "2025-05-01",
    enrolled_value: str = "yes",
    col_is_enrolled: str = "is_enrolled",
    col_nbr_visit_succeed: str = "nbr_visit_succeed",
    col_enrol_date: str = "enrollement_date_de_visite",
    col_admission_date: str = "date_admission",
    coerce_dates: bool = True
) -> pd.DataFrame:
    """
    Filtre les lignes selon la logique:
      (is_enrolled == 'yes' OR nbr_visit_succeed > 0)
      AND
      (enrollement_date_de_visite >= date_threshold OR date_admission >= date_threshold)

    Args:
        df: DataFrame d'entrée.
        date_threshold: Date seuil au format 'YYYY-MM-DD'.
        enrolled_value: Valeur considérée comme 'inscrit'.
        col_*: Noms des colonnes (si différents dans df).
        coerce_dates: Si True, convertit les colonnes de dates en datetime (errors='coerce').

    Returns:
        Un DataFrame filtré.
    """
    # Copie légère pour éviter de modifier df en place
    data = df.copy()

    # Conversion des dates si demandé
    if coerce_dates:
        for c in (col_enrol_date, col_admission_date):
            if c in data.columns:
                data[c] = pd.to_datetime(data[c], errors="coerce")

    # Date seuil en datetime
    dt = pd.to_datetime(date_threshold)

    # Sécurisation des colonnes manquantes
    if col_is_enrolled not in data.columns:
        data[col_is_enrolled] = pd.NA
    if col_nbr_visit_succeed not in data.columns:
        data[col_nbr_visit_succeed] = 0
    if col_enrol_date not in data.columns:
        data[col_enrol_date] = pd.NaT
    if col_admission_date not in data.columns:
        data[col_admission_date] = pd.NaT

    # Convert nbr_visit_succeed to numeric
    if col_nbr_visit_succeed in data.columns:
        data[col_nbr_visit_succeed] = pd.to_numeric(data[col_nbr_visit_succeed], errors='coerce')

    # Condition booléenne
    cond_left = (data[col_is_enrolled] == enrolled_value) | (data[col_nbr_visit_succeed].fillna(0) > 0)
    cond_right = (data[col_enrol_date] >= dt) | (data[col_admission_date] >= dt)
    condition = cond_left & cond_right

    return data[condition]
#=======================================================================================================================

#Function to get the age range
#Function to get the age range
def age_range(age):
    """
    Catégorise l'âge en mois en tranches d'âge pour l'analyse nutritionnelle
    
    Args:
        age (int/float): âge en mois
        
    Returns:
        str: catégorie d'âge
    """
    # Gérer les valeurs NaN ou négatives
    if pd.isna(age) or age < 0:
        return 'Age_non_defini'
    
    if age < 6:
        return 'Enfants<6mois'
    elif 6 <= age <= 23:  # Correction: utiliser <= au lieu de range()
        return 'Enfants_6_23_mois'
    elif 24 <= age <= 66:  # Correction: utiliser <= au lieu de range()
        return 'Enfants_24_59_mois'
    else:
        return 'Enfants_+59_mois'

#OVC['age_range'] = OVC['age'].map(age_range)
#OVC1 = OVC[OVC['age_range'] != '21+']
#==================================================================================
#========================================================================================================================
today_str = datetime.today().strftime('%Y-%m-%d')
enroled = pd.read_excel(f"~/Downloads/caris-meal-app/data/Nutrition (created 2025-04-25) {today_str}.xlsx",
                          parse_dates=True)
enroled_col = [
    "caseid", "name", "eligible", "manutrition_type", "date_of_birth",
    "gender", "muac", "nbr_visit", "is_alive", "death_date",
    "death_reason", "nbr_visit_succeed", "admission_muac", "office", "commune",
    "departement", "household_collection_date", "household_number", "has_household", "closed",
    "closed_date", "last_modified_date", "opened_date", "case_link", "enrollement_date_de_visite",
    "enrollment_date", "enrollment_eligibility", "enrollment_manutrition_type", "is_enrolled", "hiv_test_done",
    "hiv_test_result", "club_id", "club_name","date_admission"
]
enroled = enroled[enroled_col]
print(f"Fichier enrollement Télechargé avec {enroled.shape[0]} lignes")
#=======================================================================================================================
depistage = pd.read_excel(f"~/Downloads/caris-meal-app/data/Caris Health Agent - NUTRITON[HIDDEN] - Dépistage Nutritionnel (created 2025-06-26) {today_str}.xlsx",
                          parse_dates=True)
dep_col = [
    "form.depistage.date_de_visite", "form.depistage.last_name", "form.depistage.first_name", "form.depistage.gender",
    "form.depistage.date_of_birth", "form.depistage.muac", "form.depistage.weight_kg", "form.depistage.height",
    
    "form.depistage.edema", "form.depistage.lesion_cutane", "form.depistage.diarrhea", "form.depistage.autres_symptomes",
    "form.depistage.phone_number", "form.depistage.photo_depistage", "form.depistage.office", "form.depistage.departement",
    
    "form.depistage.commune", "form.depistage.date_de_depistage", "form.depistage.fullname", "form.depistage.eligible",
    "form.depistage.manutrition_type", "form.case.@case_id", "completed_time", "started_time",
    
    "username", "received_on", "form_link"
]
depistage = depistage[dep_col]
print(f"dépistage télechargés avec {depistage.shape[0]} lignes")
#========================================================================================================================
#Caris Health Agent - NUTRITON[HIDDEN] - Dépistage Nutritionnel (created 2025-06-26) 2025-09-23

depistage_clean = clean_column_names(depistage, expr_to_remove='form.depistage.')
end_date_week = pd.to_datetime('today')
# Date de début = 7 jours avant
start_date_week = end_date_week - timedelta(days=7)

start_date = pd.Timestamp("2025-05-01")
end_date = pd.Timestamp(datetime.today().date())

print("=== Nombre depistage pour la semaine ===")
depistage_week = extraire_data(df=depistage_clean, start_date=start_date_week, end_date=end_date_week, date_col='date_de_visite')
print(f"{depistage_week.shape[0]} dépistage réalisés pour la semaine")
print("=== Nombre depistage de mai 2025 à aujourd'hui ===")
depistage_nut = extraire_data(df=depistage_clean, start_date=start_date, end_date=end_date, date_col='date_de_visite')
print(f"{depistage_nut.shape[0]} dépistage réalisés pour la periode")
depistage_nut.to_excel("depistage_nutritionel.xlsx", sheet_name="Mai_a_aujourdhui", index=False)
#===============================================================================================
depistage_index=create_binary_symptom_columns(depistage_nut, 'autres_symptomes')
depistage_index.to_excel("depistage_index.xlsx", sheet_name="index", index=False)
# Ajouter l'âge au DataFrame de dépistage
depistage_nut = get_age_in_year(depistage_nut, 'date_of_birth')
depistage_nut = get_age_in_months(depistage_nut, 'date_of_birth')
depistage_nut['age_range'] = depistage_nut['age_months'].map(age_range)
print(f"Colonne âge ajoutée. Échantillon:")
#print(depistage_nut[['date_of_birth', 'age_years']].head())
depistage_indice = create_normalized_health_index(depistage_nut)
depistage_indice.to_excel("depistage_indice.xlsx", sheet_name="indice", index=False)
#nut_filtered = filter_patients(enroled, date_threshold="2025-05-01")

enroled["enrollement_date_de_visite"] = pd.to_datetime(enroled["enrollement_date_de_visite"], errors="coerce")
enroled["date_admission"] = pd.to_datetime(enroled["date_admission"], errors="coerce")
# Définition du seuil
date_limite = pd.to_datetime("2025-05-01")
# Convert to numeric first
enroled["nbr_visit_succeed"] = pd.to_numeric(enroled["nbr_visit_succeed"], errors='coerce').fillna(0)
# Condition
condition = (
    ((enroled["is_enrolled"] == "yes") | (enroled["nbr_visit_succeed"] > 0))
    &
    ((enroled["enrollement_date_de_visite"] >= date_limite) | (enroled["date_admission"] >= date_limite))
)
# Application du filtre
nut_filtered = enroled[condition]
nut_filtered.to_excel("Nutrition_all.xlsx")
print(f"Nombre d'enrollement avec doublons possibles {nut_filtered.shape[0]} lignes")
#==========================================================================================

#========================================================================================================
# Vous pouvez aussi l'appliquer au DataFrame enroled
nut_filtered = get_age_in_year(nut_filtered, 'date_of_birth')
nut_filtered = get_age_in_months(nut_filtered, 'date_of_birth')
nut_filtered['age_range'] = nut_filtered['age_months'].map(age_range)
nut_filtered.to_excel("Enroled.xlsx", sheet_name="enroled", index=False)

#========================================================================
depistage = depistage.rename(columns={'form.case.@case_id': 'caseid','form.depistage.date_de_visite':'date_de_visite'})
nutrition = pd.merge(nut_filtered, depistage[['date_de_visite','caseid','username']])
nutrition.to_excel("nutrition.xlsx", sheet_name="enroled", index=False)

print("=== Alertes doublons ===")
import pandas as pd
from difflib import SequenceMatcher

import pandas as pd
import numpy as np
import unicodedata
from difflib import SequenceMatcher
from typing import List

def _normalize_text_series(s: pd.Series) -> pd.Series:
    """Minuscules, trimming, compression des espaces et suppression des accents."""
    s = s.astype(str).str.lower().str.strip().str.replace(r"\s+", " ", regex=True)
    return s.apply(lambda x: unicodedata.normalize("NFKD", x).encode("ascii","ignore").decode("ascii"))

def _similar(a: str, b: str) -> float:
    """Similarité [0,100] via difflib (rapide, standard)."""
    return 100.0 * SequenceMatcher(None, a, b).ratio()

def detecter_doublons_avec_groupes(
    df: pd.DataFrame,
    colonnes: List[str],
    threshold: int = 100,
    return_only_duplicates: bool = True,
    keep_most_na: bool = False  # ✅ Nouveau paramètre
) -> pd.DataFrame:
    """
    Détecte des doublons (stricts si threshold=100, sinon fuzzy) sur `colonnes`,
    et retourne un DataFrame avec:
      - duplicate_group_id (int >= 1 pour les groupes; 0 si singleton)
      - duplicate_group_size (taille du groupe)
    
    Si keep_most_na=True, garde l'enregistrement avec le plus de valeurs N/A par groupe.
    
    Args
    ----
    df : DataFrame source
    colonnes : colonnes utilisées pour la comparaison
    threshold : 100 => exact (sur texte normalisé) ; <100 => fuzzy
    return_only_duplicates : si True, ne retourne que les lignes appartenant
                             à un groupe de taille >= 2
    keep_most_na : si True, garde l'enregistrement avec le plus de N/A par groupe
    
    Returns
    -------
    DataFrame enrichi (ou nettoyé si keep_most_na=True)
    """
    if not colonnes:
        raise ValueError("Aucune colonne fournie.")
    manquantes = [c for c in colonnes if c not in df.columns]
    if manquantes:
        raise ValueError(f"Colonnes manquantes dans df: {manquantes}")

    n = len(df)
    if n == 0:
        out = df.copy()
        out["duplicate_group_id"] = pd.Series(dtype=int)
        out["duplicate_group_size"] = pd.Series(dtype=int)
        return out

    # 1) Normalisation des colonnes de comparaison
    df_cmp = df[colonnes].copy()
    for c in colonnes:
        df_cmp[c] = _normalize_text_series(df_cmp[c])

    # 2) Cas strict (exact sur normalisé) : équivaut à un groupby sur les colonnes normalisées
    if threshold >= 100:
        # clé exacte
        keys = df_cmp.apply(lambda r: tuple(r.values.tolist()), axis=1)
        # group id via factorize
        codes, uniques = pd.factorize(keys, sort=True)
        # taille par groupe
        sizes = pd.Series(codes).map(pd.Series(pd.Series(codes).value_counts()))
        group_id = np.where(sizes.values >= 2, codes + 1, 0)  # 0 pour singletons
        out = df.copy()
        out["duplicate_group_id"] = group_id
        out["duplicate_group_size"] = np.where(group_id > 0, sizes.values, 1)
        
        # ✅ Traitement keep_most_na si demandé
        if keep_most_na:
            return _process_keep_most_na(out, return_only_duplicates)
        
        return out[out["duplicate_group_id"] > 0].reset_index(drop=True) if return_only_duplicates else out

    # 3) Cas fuzzy : union-find sur paires similaires (>= threshold)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    # Comparaison naïve toutes paires (optimisez avec un blocking si besoin)
    vals = df_cmp.values.astype(str)
    for i in range(n):
        vi = vals[i]
        for j in range(i+1, n):
            vj = vals[j]
            # score minimal sur les colonnes (toutes doivent dépasser le seuil)
            smin = 100.0
            for k in range(len(colonnes)):
                s = _similar(vi[k], vj[k])
                if s < smin:
                    smin = s
                if smin < threshold:
                    break
            if smin >= threshold:
                union(i, j)

    # Construire les groupes
    roots = [find(i) for i in range(n)]
    # remap racines -> id 1..G pour lisibilité
    uniq_roots = {}
    next_id = 1
    group_id = []
    for r in roots:
        if r not in uniq_roots:
            uniq_roots[r] = next_id
            next_id += 1
        group_id.append(uniq_roots[r])

    # Taille des groupes
    counts = pd.Series(group_id).value_counts()
    size_map = counts.to_dict()
    size_vec = [size_map[g] for g in group_id]

    # Marquer les singletons comme 0
    group_id_final = [gid if size_map[gid] >= 2 else 0 for gid in group_id]
    size_final = [size_map[gid] if size_map[gid] >= 2 else 1 for gid in group_id]

    out = df.copy()
    out["duplicate_group_id"] = group_id_final
    out["duplicate_group_size"] = size_final

    # ✅ Traitement keep_most_na si demandé
    if keep_most_na:
        return _process_keep_most_na(out, return_only_duplicates)

    return out[out["duplicate_group_id"] > 0].reset_index(drop=True) if return_only_duplicates else out

def _process_keep_most_na(df_with_groups: pd.DataFrame, return_only_duplicates: bool = True) -> pd.DataFrame:
    """
    Fonction auxiliaire pour traiter les groupes et garder l'enregistrement avec le plus de N/A.
    """
    
    # 1) Calculer le pourcentage de N/A pour chaque ligne
    # Exclure les colonnes techniques ajoutées par la détection
    technical_cols = ['duplicate_group_id', 'duplicate_group_size']
    data_cols = [col for col in df_with_groups.columns if col not in technical_cols]
    
    total_cols = len(data_cols)
    df_with_groups['na_count'] = df_with_groups[data_cols].isna().sum(axis=1)
    df_with_groups['na_percentage'] = (df_with_groups['na_count'] / total_cols * 100).round(2)
    
    print(f"📊 Pourcentage de N/A calculé sur {total_cols} colonnes")
    print(f"📈 Statistiques N/A - Min: {df_with_groups['na_percentage'].min()}%, Max: {df_with_groups['na_percentage'].max()}%")
    
    # 2) Fonction pour garder l'enregistrement avec le plus de N/A par groupe
    def keep_most_na_per_group(group):
        if len(group) == 1:
            return group  # Pas de doublon, garder tel quel
        
        # Trier par pourcentage N/A décroissant, puis par na_count pour tie-breaking
        group_sorted = group.sort_values(['na_percentage', 'na_count'], ascending=[False, False])
        most_na = group_sorted.iloc[0:1]  # Garder le premier (plus de N/A)
        
        # Afficher uniquement pour les vrais groupes de doublons
        if 'caseid' in group.columns:
            print(f"   Groupe {group['duplicate_group_id'].iloc[0]}: {len(group)} doublons → gardé caseid {most_na['caseid'].iloc[0]} ({most_na['na_percentage'].iloc[0]}% N/A)")
        else:
            print(f"   Groupe {group['duplicate_group_id'].iloc[0]}: {len(group)} doublons → gardé index {most_na.index[0]} ({most_na['na_percentage'].iloc[0]}% N/A)")
        
        return most_na
    
    # 3) Traiter les groupes
    print("🔍 Traitement des groupes de doublons...")
    
    # Séparer les non-doublons (duplicate_group_id == 0) et les doublons
    non_duplicates = df_with_groups[df_with_groups['duplicate_group_id'] == 0].copy()
    duplicates = df_with_groups[df_with_groups['duplicate_group_id'] > 0].copy()
    
    print(f"   📋 {len(non_duplicates)} enregistrements uniques")
    print(f"   🔍 {len(duplicates)} enregistrements en doublons dans {duplicates['duplicate_group_id'].nunique()} groupes")
    
    # Appliquer la fonction de sélection sur chaque groupe de doublons
    if len(duplicates) > 0:
        kept_duplicates = duplicates.groupby('duplicate_group_id').apply(keep_most_na_per_group).reset_index(drop=True)
    else:
        kept_duplicates = pd.DataFrame()
    
    # 4) Combiner les résultats
    result = pd.concat([non_duplicates, kept_duplicates], ignore_index=True)
    
    # 5) Nettoyer les duplicate_group_id pour qu'ils soient uniques
    # Réinitialiser duplicate_group_id à 0 pour tous (plus de doublons maintenant)
    result['duplicate_group_id'] = 0
    result['duplicate_group_size'] = 1
    
    # 6) Supprimer les colonnes de travail
    result = result.drop(['na_count', 'na_percentage'], axis=1)
    
    # 7) Vérification finale
    print(f"✅ Résultat final: {len(result)} enregistrements")
    print(f"🎯 Tous les duplicate_group_id sont à 0 (uniques): {(result['duplicate_group_id'] == 0).all()}")
    
    # Vérifier l'unicité des caseid si la colonne existe
    if 'caseid' in result.columns:
        caseid_duplicates = result['caseid'].duplicated().sum()
        if caseid_duplicates > 0:
            print(f"⚠️ Attention: {caseid_duplicates} caseid encore en doublon")
            # Supprimer les doublons restants par caseid (garder le premier)
            result = result.drop_duplicates(subset=['caseid'], keep='first')
            print(f"🧹 Après nettoyage final: {len(result)} enregistrements")
        else:
            print(f"✅ Tous les caseid sont uniques")
    
    return result

# ✅ REMPLACEMENT DE VOTRE CODE EXISTANT (lignes 551-565)
print("\n=== TRAITEMENT AVANCÉ DES DOUBLONS AVEC FONCTION INTÉGRÉE ===")

# Appliquer la détection avec traitement automatique des N/A
nutrition_clean = detecter_doublons_avec_groupes(
    nutrition, 
    colonnes=["name", "commune", "username"], 
    threshold=90,
    return_only_duplicates=False,  # Garder tous les enregistrements
    keep_most_na=True  # ✅ Activer le traitement N/A intégré
)

# Sauvegarder le résultat (A corriger en automatique sur muac)
nutrition_clean['manutrition_type'] = (
    nutrition_clean['manutrition_type']
    .replace('---', 'MAM')
    .fillna('MAM')
)
nutrition_clean.to_excel("nutrition_sans_doublon_integrated.xlsx", index=False)

# ✅ Vérification supplémentaire avec isin() pour s'assurer de la suppression
print("\n🔍 Vérification avec isin()...")

# Obtenir les caseid gardés
if 'caseid' in nutrition_clean.columns:
    kept_caseids = nutrition_clean['caseid'].tolist()
    
    # Vérifier que tous les caseid du résultat existent bien dans le DataFrame original
    original_caseids = set(nutrition['caseid'].tolist())
    kept_caseids_set = set(kept_caseids)
    removed_caseids = original_caseids - kept_caseids_set
    
    print(f"📊 Résumé du nettoyage:")
    print(f"   - DataFrame original: {len(nutrition)} enregistrements")
    print(f"   - DataFrame nettoyé: {len(nutrition_clean)} enregistrements")
    print(f"   - Enregistrements supprimés: {len(removed_caseids)}")
    print(f"   - Taux de suppression: {len(removed_caseids)/len(nutrition)*100:.1f}%")
    
    # Validation finale avec isin()
    validation_mask = nutrition['caseid'].isin(kept_caseids)
    validation_df = nutrition[validation_mask]
    
    if len(validation_df) == len(nutrition_clean):
        print("✅ Validation isin() réussie: le résultat est cohérent")
    else:
        print(f"⚠️ Incohérence détectée: {len(validation_df)} vs {len(nutrition_clean)}")
    
    # Sauvegarder aussi la liste des caseid supprimés pour audit
    if removed_caseids:
        removed_df = pd.DataFrame({'removed_caseid': list(removed_caseids)})
        removed_df.to_excel("caseids_supprimes_integrated.xlsx", index=False)

# ✅ BONUS: Analyse des N/A par colonne
print("\n📈 Analyse des valeurs manquantes par colonne dans le résultat final:")
na_analysis = nutrition_clean.isnull().sum().sort_values(ascending=False)
na_percentage_by_col = (na_analysis / len(nutrition_clean) * 100).round(1)

for col, count in na_analysis.head(10).items():
    print(f"   {col}: {count} N/A ({na_percentage_by_col[col]}%)")

# Mise à jour de la variable nutrition pour la suite du pipeline
#nutrition = nutrition_clean.copy()
#==============================================================
# 1) Doublons STRICTS (casse/accents/espaces ignorés)
res_fuzzy = detecter_doublons_avec_groupes(nutrition, colonnes=["name","commune","username","date_of_birth"], threshold=100)
res_fuzzy.to_excel("doublon_nut_strict.xlsx")

# 2) Doublons FUZZY (tolère petites fautes/variantes)
nut_fuzzy = detecter_doublons_avec_groupes(nutrition, colonnes=["name","commune","username"], threshold=90)
nut_fuzzy.to_excel("doublon_nut_fuzzy.xlsx")

# Exemple : on suppose que df contient déjà les colonnes citées
#============================================================================================
# ✅ CORRECTION DU PIVOT DEPISTAGE AVEC DEPARTEMENT
print("\n=== CREATION PIVOT DEPISTAGE AVEC DEPARTEMENT ===")

# Pivot table : une ligne par Département-Commune-Âge (en mois), colonnes = catégories (MAS, MAM, Normal…)
depistage_nut = depistage_nut.rename(columns={'form.case.@case_id': 'caseid','form.depistage.date_de_visite':'date_de_visite'})
matrix_depistage = depistage_nut[["commune","age_range","manutrition_type","caseid"]].copy()
depistage_nut['manutrition_type'] = depistage_nut['manutrition_type'].fillna('Normal')

# Catégories ordonnées
depistage_nut['manutrition_type'] = pd.Categorical(
    depistage_nut['manutrition_type'], 
    categories=['MAM','MAS', 'Normal'], 
    ordered=True
)

depistage_nut['age_range'] = pd.Categorical(
    depistage_nut['age_range'], 
    categories=['Enfants<6mois', 'Enfants_6_23_mois', 'Enfants_24_59_mois'], 
    ordered=True
)

# Pivot SANS reset_index() d'abord
pivot_depistage = pd.pivot_table(
    depistage_nut,
    index=["commune"],  # commune reste l'index
    columns=["age_range","manutrition_type"],        
    values="caseid",            
    aggfunc="count",             
    fill_value=0               
)
# ⚠️ Ne pas faire reset_index() tout de suite

print(f"📊 Pivot créé avec {len(pivot_depistage)} communes")

# Construire le mapping unique commune -> departement (Series)
dep_map = (
    depistage_nut.loc[:, ["commune","departement"]]
            .dropna(subset=["commune"])
            .drop_duplicates(subset=["commune"])
            .set_index("commune")["departement"]
)

print(f"📋 Mapping département créé pour {len(dep_map)} communes")

# ✅ Ajouter 'departement' au pivot en utilisant l'index 'commune'
pivot_depistage_with_dep = pivot_depistage.copy()
pivot_depistage_with_dep.insert(0, "departement", pivot_depistage_with_dep.index.map(dep_map))

# ✅ MAINTENANT faire reset_index et ordonner
pivot_depistage_with_dep = pivot_depistage_with_dep.reset_index().sort_values(['departement', 'commune'])

pivot_depistage_with_dep.to_excel("depistage_table.xlsx")
print("💾 Fichier sauvegardé: depistage_table.xlsx")
#=======================================================================================================================

# 1) Sous-ensemble et catégories (garde ce que tu avais)
matrix_nutrition = nutrition_clean[["commune","age_range","manutrition_type","caseid"]].copy()
matrix_nutrition["manutrition_type"] = pd.Categorical(
    matrix_nutrition["manutrition_type"].str.strip(),
    categories=["MAS","MAM","Normal"], ordered=True
)
matrix_nutrition["age_range"] = pd.Categorical(
    matrix_nutrition["age_range"].str.strip(),
    categories=["Enfants<6mois","Enfants_6_23_mois","Enfants_24_59_mois"],
    ordered=True
)

# 2) Pivot "pur": index = commune ; colonnes = (age_range, manutrition_type)
pivot_enroled = pd.pivot_table(
    matrix_nutrition,
    index="commune",
    columns=["manutrition_type","age_range"],
    values="caseid",
    aggfunc="count",
    fill_value=0
)

# 3) Construire un mapping unique commune -> departement (Series)
dep_map = (
    nutrition.loc[:, ["commune","departement"]]
            .dropna(subset=["commune"])
            .drop_duplicates(subset=["commune"])
            .set_index("commune")["departement"]
)

# 4) Ajouter 'departement' au pivot par mapping d'index (pas de merge)
pivot_enroled_with_dep = pivot_enroled.copy()
pivot_enroled_with_dep.insert(0, "departement", pivot_enroled_with_dep.index.map(dep_map))
# ✅ Reset index et ordonner par département puis commune
pivot_enroled_with_dep = pivot_enroled_with_dep.reset_index().sort_values(['departement', 'commune'])

# 5) Export
pivot_enroled_with_dep.to_excel("enrol_table.xlsx")
#=======================================================================================================================
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from pathlib import Path

def copy_ranges_by_mapping(
    mapping: dict,
    rows_source_range: str,
    src_file: str = "enrol_table.xlsx",
    dst_file: str = "cluster.xlsx",
    src_sheet: str | None = None,
    dst_sheet: str | None = None,
):
    """
    Copie des données par blocs de cellules d'un fichier Excel source vers un fichier destination.

    Args:
        mapping: dict {"SRC_RANGE": "DST_RANGE", ...}
                 Ex: {"C4:C22": "A4:A22", "B4:B22": "B4:B22", "D4:F22": "C4:C22"}
        rows_source_range: plage servant à calculer le nombre de lignes à copier (ex: "B4:B22")
        src_file: Excel source (par défaut "enrol_table.xlsx")
        dst_file: Excel destination (par défaut "cluster.xlsx")
        src_sheet: nom de feuille source (None = active)
        dst_sheet: nom de feuille destination (None = active)

    Règles:
      - La hauteur copiée = nombre de lignes de rows_source_range (ou nombre de cellules non vides si plus court).
      - Si largeur src == largeur dst -> copie bloc à bloc.
      - Si largeur src > 1 et largeur dst == 1 -> somme horizontale par ligne (collapsing).
      - Sinon -> erreur.
    """
    # Ouvrir classeurs
    wb_src = load_workbook(src_file, data_only=True)
    wb_dst = load_workbook(dst_file)

    ws_src = wb_src[src_sheet] if src_sheet else wb_src.active
    ws_dst = wb_dst[dst_sheet] if dst_sheet else wb_dst.active

    # -------------------------
    # 1) Déterminer n_rows
    # -------------------------
    rs_min_col, rs_min_row, rs_max_col, rs_max_row = range_boundaries(rows_source_range)
    # hauteur théorique
    n_rows = rs_max_row - rs_min_row + 1
    # compter non vides si tu veux tronquer aux non vides en tête
    non_empty = 0
    for r in range(rs_min_row, rs_min_row + n_rows):
        v = ws_src.cell(row=r, column=rs_min_col).value
        if v is not None and (not isinstance(v, str) or v.strip() != ""):
            non_empty += 1
        else:
            # si tu préfères t'arrêter au premier vide, décommente :
            # break
            pass
    # Choix: on garde la hauteur fournie par la plage; si tu veux limiter aux non vides, utilise:
    # n_rows = max(non_empty, 0)

    # -------------------------
    # 2) Boucle de copie
    # -------------------------
    for src_rng, dst_rng in mapping.items():
        smin_c, smin_r, smax_c, smax_r = range_boundaries(src_rng)
        dmin_c, dmin_r, dmax_c, dmax_r = range_boundaries(dst_rng)

        src_width = smax_c - smin_c + 1
        dst_width = dmax_c - dmin_c + 1

        # Vérifs de cohérence hauteur
        dst_height = dmax_r - dmin_r + 1
        if dst_height < n_rows:
            raise ValueError(
                f"Destination {dst_rng} trop courte ({dst_height} lignes) pour copier {n_rows} lignes."
            )

        if src_width == dst_width:
            # Copie 1:1 des blocs
            for rr in range(n_rows):
                for cc in range(src_width):
                    val = ws_src.cell(row=smin_r + rr, column=smin_c + cc).value
                    ws_dst.cell(row=dmin_r + rr, column=dmin_c + cc, value=val)

        elif src_width > 1 and dst_width == 1:
            # Cas collapsing: on somme horizontalement la source pour déposer dans 1 colonne
            for rr in range(n_rows):
                row_sum = 0
                for cc in range(src_width):
                    v = ws_src.cell(row=smin_r + rr, column=smin_c + cc).value
                    if v is None or (isinstance(v, str) and v.strip() == ""):
                        continue
                    try:
                        row_sum += float(v)
                    except Exception:
                        # Si non numérique, on ignore (ou lève une erreur selon tes besoins)
                        pass
                ws_dst.cell(row=dmin_r + rr, column=dmin_c, value=row_sum)
        else:
            raise ValueError(
                f"Incompatibilité largeur: source {src_rng} (w={src_width}) -> destination {dst_rng} (w={dst_width})."
            )

    # Enregistrer
    out_path = Path(dst_file).with_name(Path(dst_file).stem + "_filled.xlsx")
    wb_dst.save(out_path)
    return str(out_path)


# -----------------------------
# Exemple d'utilisation
# -----------------------------
if __name__ == "__main__":
    # Mapping souhaité:
    #  - enrole1 = C4:C22 -> cluster1 = A4:A22
    #  - enrole2 = B4:B22 -> cluster2 = B4:B22
    #  - enrole3 = D4:F22 -> cluster3 = C4:C22  (somme des 3 colonnes vers 1 colonne)
    mapping = {
        "C4:C22": "A4:A22",
        "B4:B22": "B4:B22",
        "D4:F22": "C4:E22",
    }

    # La hauteur de copie se base sur enrole2 = B4:B22
    rows_ref = "B4:B22"

    out = copy_ranges_by_mapping(
        mapping=mapping,
        rows_source_range=rows_ref,
        src_file="enrol_table.xlsx",
        dst_file="cluster.xlsx",
        src_sheet=None,   # ou "Feuil1"
        dst_sheet=None    # ou "cluster"
    )
    print("Fichier généré:", out)

print("\n" + "="*60)
print("FIN DU SCRIPT AVEC COPIE CORRIGÉE")
print("="*60)