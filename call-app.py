#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
CallApp Analysis Script
Converted from Jupyter Notebook to executable Python script
"""

# Importing packages

import matplotlib.pyplot as plt
import plotly.express as px
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Fill, PatternFill, Protection
from datetime import datetime
import locale
import pymysql
from sqlalchemy import create_engine
import warnings
warnings.filterwarnings('ignore')
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils import get_column_letter

# Define datetime range
start_date = pd.to_datetime('2025-08-25')
end_date = pd.to_datetime('2025-08-31')  # Fixed to today's date
today_date = datetime.today().date().strftime('%Y-%m-%d')
mois_ordre = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

def transform_to_month_year_french(df, date_column):
    if date_column not in df.columns:
        raise ValueError(f"La colonne '{date_column}' n'existe pas dans le DataFrame.")
    df[date_column] = pd.to_datetime(df[date_column])
    df['mois'] = df[date_column].dt.strftime('%B').str.capitalize()
    df['annee'] = df[date_column].dt.strftime('%Y').str.capitalize()
    return df

def create_pivot_table(df, oev_patient_code):
    # Création du tableau croisé dynamique (pivot table)
    pivot_table = pd.pivot_table(
        df,
        values=oev_patient_code,
        index=['mois'],
        columns=['commune', 'Trouvé'],
        aggfunc=lambda x: len(x),
        margins=True,
        margins_name='Total'
    ).fillna(0)

    mois_ordre = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    pivot_table = pivot_table.reindex(mois_ordre).fillna(0)
    
    # Filtrage des lignes où toutes les valeurs ne sont pas égales à zéro
    filtered_pivot_table = pivot_table[(pivot_table != 0).any(axis=1)]
    
    return filtered_pivot_table

def update_column(df, search_column, search_value, is_column, is_value):
    """
    Met à jour la colonne `is_ugp_column` à `is_ugp_value` pour les lignes où 
    `search_column` est égal à `search_value`.

    :param df: DataFrame à modifier
    :param search_column: Le nom de la colonne a selectionner
    :param search_value: La valeur de la colonne à filtrer
    :param is_column: Le nom de la colonne à mettre à jour
    :param is_value: La valeur à attribuer à la colonne mise à jour
    """
    df.loc[(df[search_column] == search_value), is_column] = is_value

def process_grouped(df, date_col, value_col, group_col):
    """
    Regroupe un DataFrame par semaine (lundi début) et par programme,
    puis crée un pivot table avec les valeurs agrégées.

    :param df: DataFrame contenant les données
    :param date_col: Nom de la colonne contenant les dates
    :param value_col: Nom de la colonne contenant les valeurs à compter/sommer
    :param group_col: Nom de la colonne contenant les groupes (par exemple, 'Programme')
    :return: DataFrame pivoté
    """
    # Convertir la colonne de date en datetime
    df[date_col] = pd.to_datetime(df[date_col])

    # Calculer la date du début de chaque semaine (lundi) et supprimer l'heure
    df['Semaine'] = (df[date_col] - pd.to_timedelta(df[date_col].dt.weekday, unit='D')).dt.date

    # Créer un pivot table pour regrouper les données
    df_pivot = df.pivot_table(
        index="Semaine",          # Regroupement par semaine
        columns=group_col,        # Séparation par programme
        values=value_col,         # Somme des valeurs à compter
        aggfunc="sum",            # Fonction d'agrégation (somme)
        fill_value=0,             # Remplacer NaN par 0
        margins = True,
        margins_name="Total"  # Nom du total pour les lignes et colonnes marginales
        
    )

    # Réinitialiser l'index pour obtenir un DataFrame propre
    df_pivot = df_pivot.reset_index()

    return df_pivot

def export_multiple_pivot_tables_to_excel(pivot_tables, file_name, sheet_names, titles, title_font_size=14, title_alignment='center', delete_row=None):
    # Création du fichier Excel avec ordre spécifique pour "Tableau des Appels et Visites"
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        # Traiter les feuilles dans l'ordre spécifié pour garantir que "Tableau des Appels et Visites" soit en premier
        for i, (pivot_table, sheet_name, title) in enumerate(zip(pivot_tables, sheet_names, titles)):
            pivot_table.to_excel(writer, sheet_name=sheet_name, startrow=2)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Ajouter un titre
            cell = worksheet.cell(row=1, column=1)
            cell.value = title
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
            cell.font = Font(bold=True, size=title_font_size)
            cell.alignment = Alignment(horizontal=title_alignment)

            # Ajuster la largeur des colonnes
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column  # Get the column index
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)

            # Supprimer une ligne spécifique si nécessaire
            if delete_row:
                worksheet.delete_rows(delete_row)
        
        # S'assurer que "Tableau des Appels et Visites" est en première position
        # et que "Listes des Appels et Visites" est après
        if "Tableau des Appels et Visites" in sheet_names:
            workbook = writer.book
            all_sheets = list(workbook.sheetnames)
            
            # Placer "Tableau des Appels et Visites" en première position
            if "Tableau des Appels et Visites" in all_sheets and all_sheets[0] != "Tableau des Appels et Visites":
                target_sheet = workbook["Tableau des Appels et Visites"]
                workbook.move_sheet(target_sheet, offset=-len(all_sheets))
                print(f"[OK] Feuille 'Tableau des Appels et Visites' placée en première position dans le fichier Excel.")
            
            # S'assurer que "Listes des Appels et Visites" est après "Tableau des Appels et Visites"
            updated_sheets = list(workbook.sheetnames)
            if "Listes des Appels et Visites" in updated_sheets:
                tableau_index = updated_sheets.index("Tableau des Appels et Visites") if "Tableau des Appels et Visites" in updated_sheets else -1
                liste_index = updated_sheets.index("Listes des Appels et Visites")
                
                if tableau_index >= 0 and liste_index < tableau_index:
                    # Déplacer "Listes des Appels et Visites" juste après "Tableau des Appels et Visites"
                    liste_sheet = workbook["Listes des Appels et Visites"]
                    workbook.move_sheet(liste_sheet, offset=(tableau_index - liste_index))
                    print(f"[OK] Feuille 'Listes des Appels et Visites' placée après 'Tableau des Appels et Visites'.")
                elif tableau_index >= 0 and liste_index > tableau_index:
                    print(f"[OK] L'ordre correct est déjà respecté: 'Tableau des Appels et Visites' avant 'Listes des Appels et Visites'.")

    print(f"Les tableaux croisés dynamiques ont été exportés vers '{file_name}'.")

def export_pivot_table_to_excel(pivot_table, file_name, sheet_name, title, title_font_size=14, title_alignment='center', delete_row=None):
    # Création du fichier Excel
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        pivot_table.to_excel(writer, sheet_name=sheet_name, startrow=2)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Ajouter un titre
        cell = worksheet.cell(row=1, column=1)
        cell.value = title
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        cell.font = Font(bold=True, size=title_font_size)
        cell.alignment = Alignment(horizontal=title_alignment)

        # Ajuster la largeur des colonnes
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column index
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)

        # Supprimer une ligne spécifique si nécessaire
        if delete_row:
            worksheet.delete_rows(delete_row)

    print(f"Le tableau croisé dynamique a été exporté vers '{file_name}'.")

# Function to assign commune names using regex
def assign_commune(name):
    if re.match(r'^1', name):
        return 'Cap-Haïtien'
    elif re.match(r'^2', name):
        return 'Port-au-Prince'
    elif re.match(r'^5', name):
        return 'Port-de-Paix'
    elif re.match(r'^6', name):
        return 'Gonaïves'
    else:
        return 'Autre'

# Modify the dataframe_for_period function to accept start_date and end_date as parameters
def dataframe_for_period(df, date_column, start_date=None, end_date=None):
    """
    Filter dataframe by date range and assign commune based on username
    """
    if start_date is None:
        start_date = pd.to_datetime('2025-08-25')
    if end_date is None:
        end_date = pd.to_datetime('2025-08-31')
    
    df[date_column] = pd.to_datetime(df[date_column])
    # Filter by date range
    df = df[(df[date_column] >= start_date) & (df[date_column] <= end_date)]

    # Assign 'commune' based on 'username'
    df['commune'] = df['username'].apply(assign_commune)

    return df

def add_found_percentage(data_pivotable):
    """
    Ajoute une colonne `Total_Trouvé` (somme des 'Oui' au niveau 1) et une colonne 
    `% Trouvé` (ratio de 'Total_Trouvé' par rapport à 'Total'), formatée en pourcentage.

    Parameters:
        data_pivotable (pd.DataFrame): Le DataFrame pivoté à modifier.

    Returns:
        pd.DataFrame: Le DataFrame mis à jour avec les colonnes supplémentaires.
    """
    # Vérifier si les colonnes nécessaires existent
    if 'Total' not in data_pivotable.columns:
        raise KeyError("La colonne 'Total' est manquante dans le DataFrame.")
    
    # Ajouter la colonne Total_Trouvé
    data_pivotable['Total_Trouvé'] = data_pivotable.xs('Oui', level=1, axis=1).sum(axis=1)

    # Ajouter la colonne % Trouvé
    data_pivotable['% Trouvé'] = data_pivotable['Total_Trouvé'] / data_pivotable['Total']

    # Formater la colonne % Trouvé en pourcentage
    data_pivotable['% Trouvé'] = data_pivotable['% Trouvé'].apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "0.00%")

    # Retourner le DataFrame mis à jour
    return data_pivotable

def add_dataframe_to_workbook(file_path, sheet_name, dataframe):
    """
    Adds a DataFrame to an existing Excel workbook in the specified sheet.

    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet where the DataFrame will be written.
        dataframe (pd.DataFrame): The DataFrame to write to the sheet.
    """
    from openpyxl import load_workbook
    
    # Load the existing workbook
    workbook = load_workbook(file_path)

    # Check if the sheet already exists, and create it if not
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
        workbook.save(file_path)

    # Write the DataFrame to the specified sheet
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"The DataFrame has been written to the sheet '{sheet_name}' in '{file_path}'.")

def generate_excel_from_dataframe(df, file_name):
    from openpyxl.styles import PatternFill, Border, Side
    
    # Créez un writer Pandas Excel en utilisant openpyxl comme moteur.
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    
    # Écrivez le DataFrame dans le fichier Excel.
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    
    # Obtenez les objets workbook et worksheet d'openpyxl.
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Définition des styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    header_fill = PatternFill(fgColor="FFFF00", fill_type="solid")
    blue_border = Border(
        left=Side(style='thin', color='0000FF'),
        right=Side(style='thin', color='0000FF'),
        top=Side(style='thin', color='0000FF'),
        bottom=Side(style='thin', color='0000FF')
    )
    
    # Ajustement des largeurs des colonnes
    for col_num, column in enumerate(df.columns.values, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = blue_border
        
        max_length = max(df[column].astype(str).apply(len).max(), len(column)) + 2
        worksheet.column_dimensions[cell.column_letter].width = max_length
    
    # Appliquez la bordure bleue aux cellules de données
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = blue_border
    
    # Ajustement de la hauteur des lignes
    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 20
    
    # Sauvegarde du fichier Excel
    writer.close()

def generate_excel_from_table(df, file_name, groupby_column, table_spacing):
    from openpyxl.styles import PatternFill, Border, Side
    
    # Créez un writer Pandas Excel en utilisant openpyxl comme moteur.
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    
    # Obtenez les objets workbook et worksheet d'openpyxl.
    writer.book.create_sheet(title='Sheet1')
    worksheet = writer.book['Sheet1']
    
    # Définition des styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    header_fill = PatternFill(fgColor="FFFF00", fill_type="solid")
    blue_border = Border(
        left=Side(style='thin', color='0000FF'),
        right=Side(style='thin', color='0000FF'),
        top=Side(style='thin', color='0000FF'),
        bottom=Side(style='thin', color='0000FF')
    )
    
    # Initialisation de la position verticale
    start_row = 1
    
    grouped_df = df.groupby(groupby_column)
    
    for group_name, group_data in grouped_df:
        # Écrire le nom du groupe
        worksheet.cell(row=start_row, column=1, value=str(group_name)).font = Font(bold=True, color="0000FF")
        start_row += 1
        
        # Écrire l'en-tête
        for col_num, column in enumerate(df.columns.values, start=1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.value = column
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = blue_border
            
            max_length = max(group_data[column].astype(str).apply(len).max(), len(column)) + 2
            worksheet.column_dimensions[cell.column_letter].width = max_length
        
        # Écrire les données
        for row_idx, row in enumerate(group_data.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = blue_border
        
        # Mise à jour de la position pour le prochain tableau
        start_row = row_idx + table_spacing + 1
    
    # Ajustement de la hauteur des lignes
    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 20
    
    # Sauvegarde du fichier Excel
    writer._save()

def export_multiple_pivot_tables(pivot_tables, file_name, sheet_name, titles, title_font_size=14, title_alignment='center', delete_row=None):
    """
    Exporte plusieurs tableaux croisés dynamiques (pivot tables) dans une seule feuille Excel,
    espacés de deux lignes les uns des autres.
    
    :param pivot_tables: Liste des DataFrames (pivot tables)
    :param file_name: Nom du fichier Excel de sortie
    :param sheet_name: Nom de la feuille
    :param titles: Liste des titres pour chaque pivot table
    :param title_font_size: Taille de la police du titre (défaut: 14)
    :param title_alignment: Alignement du titre (défaut: 'center')
    :param delete_row: Supprime une ligne spécifique (optionnel)
    """
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet(title=sheet_name) if sheet_name not in workbook.sheetnames else writer.sheets[sheet_name]
        start_row = 1
        
        for pivot_table, title in zip(pivot_tables, titles):
            # Ajouter le titre
            cell = worksheet.cell(row=start_row, column=1)
            cell.value = title
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=pivot_table.shape[1])
            cell.font = Font(bold=True, size=title_font_size)
            cell.alignment = Alignment(horizontal=title_alignment)
            
            # Écrire la pivot_table sous le titre
            for r_idx, row in enumerate(pivot_table.itertuples(), start=start_row + 2):
                for c_idx, value in enumerate(row, start=1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
            
            start_row += len(pivot_table) + 4  # Ajouter un espace de 2 lignes après chaque pivot_table
            
        # Supprimer une ligne spécifique si nécessaire
        if delete_row:
            worksheet.delete_rows(delete_row)
    
    print(f"Les tableaux croisés dynamiques ont été exportés vers '{file_name}' dans la feuille '{sheet_name}'.")

def main():
    """Main execution function"""
    print("Starting CallApp Analysis...")
    
    # Define datetime range at the beginning of main()
    start_date = pd.to_datetime('2025-08-25')
    end_date = pd.to_datetime('2025-08-31')
    today_date = datetime.today().date().strftime('%Y-%m-%d')
    
    # Import dial dataset
    print("Reading dial datasets...")
    Apel_ptme = pd.read_excel(f"C:/Users/Moise/Downloads/caris-meal-app/data/Caris Health Agent - Femme PMTE  - APPELS PTME (created 2025-02-13) {today_date}.xlsx", parse_dates=True)
    Apel_oev = pd.read_excel(f"C:/Users/Moise/Downloads/caris-meal-app/data/Caris Health Agent - Enfant - APPELS OEV (created 2025-01-08) {today_date}.xlsx", parse_dates=True)

    # Import visit dataset
    print("Reading visit datasets...")
    Visite_ptme = pd.read_excel(f"C:/Users/Moise/Downloads/caris-meal-app/data/Caris Health Agent - Femme PMTE  - Visite PTME (created 2025-02-13) {today_date}.xlsx", parse_dates=True)
    Ration_ptme = pd.read_excel(f"C:/Users/Moise/Downloads/caris-meal-app/data/Caris Health Agent - Femme PMTE  - Ration & Autres Visites (created 2025-02-18) {today_date}.xlsx", parse_dates=True)
    Ration_oev = pd.read_excel(f"C:/Users/Moise/Downloads/caris-meal-app/data/Caris Health Agent - Enfant - Ration et autres visites (created 2022-08-29) {today_date}.xlsx", parse_dates=True)
    oev_visite = pd.read_excel(f"C:/Users/Moise/Downloads/caris-meal-app/data/Caris Health Agent - Enfant - Visite Enfant (created 2025-07-30) {today_date}.xlsx", parse_dates=True)

    # We copy ration oev file to have info on oev visit
    Visite_oev = Ration_oev.copy(deep=True)

    ### Data wrangling ###
    print("Processing data...")
    
    Apel_ptme["Programme"] = "PTME"
    Visite_ptme["Programme"] = "PTME"
    Ration_ptme["Programme"] = "PTME"
    Apel_ptme["Type"] = "Appel"
    Visite_ptme["Type"] = "Visite"
    Ration_ptme["Type"] = "Visite"

    Apel_oev["Programme"] = "OEV"
    Apel_oev["Type"] = "Appel"
    Ration_oev["Programme"] = "OEV"
    Ration_oev["Type"] = "Visite"
    oev_visite["Type"] = "Visite"
    oev_visite["Programme"] = "OEV"

    # column to use in the script
    apel_oev_column = ['formid','form.appels_oev.patient_code','form.appels_oev.date_appel','form.appels_oev.parenttuteur_trouve','username','Programme','Type']
    visite_ptme_column = ['formid','form.visite_ptme.health_id','form.visite_ptme.date_of_visit','form.visite_ptme.is_present','username','Programme','Type']
    visite_ration_ptme_column = ['formid','form.visit_ratio_and_others.patient_code','form.visit_ratio_and_others.date_of_visit','form.visit_ratio_and_others.is_benficiary_present','username','Programme','Type']
    visite_ration_oev_column = ['formid','form.visit_ratio_and_others.patient_code','form.visit_ratio_and_others.date_of_visit','form.visit_ratio_and_others.is_benficiary_present','username','Programme','Type']
    oev_visite_column = ['formid','form.visite_enfant.patient_code','form.visite_enfant.date_of_visit','form.visite_enfant.is_available_at_time_visit','username','Programme','Type']
    apel_ptme_column = ['formid','form.APPELS_PTME.patient_code','form.APPELS_PTME.date_appel','form.APPELS_PTME.is_ptme_available','username','Programme','Type']

    Apel_oev = Apel_oev[apel_oev_column]
    print(f"Apel_oev shape: {Apel_oev.shape}")
    Apel_ptme = Apel_ptme[apel_ptme_column]
    print(f"Apel_ptme shape: {Apel_ptme.shape}")
    Visite_ptme = Visite_ptme[visite_ptme_column]
    print(f"Visite_ptme shape: {Visite_ptme.shape}")
    Ration_ptme = Ration_ptme[visite_ration_ptme_column]
    print(f"Ration_ptme shape: {Ration_ptme.shape}")
    Ration_oev = Ration_oev[visite_ration_oev_column]
    print(f"Ration_oev shape: {Ration_oev.shape}")
    oev_visite = oev_visite[oev_visite_column]
    print(f"oev_visite shape: {oev_visite.shape}")

    # Rename columns
    Apel_ptme.rename(columns={'form.APPELS_PTME.patient_code':'patient_code',
                       'form.APPELS_PTME.date_appel':'date',
                       'form.APPELS_PTME.is_ptme_available':'Trouvé'}, inplace=True)

    Ration_ptme.rename(columns={'form.visit_ratio_and_others.patient_code':'patient_code',
                       'form.visit_ratio_and_others.date_of_visit':'date',
                       'form.visit_ratio_and_others.is_benficiary_present':'Trouvé'}, inplace=True)

    Visite_ptme.rename(columns={'form.visite_ptme.health_id':'patient_code',
                       'form.visite_ptme.date_of_visit':'date',
                       'form.visite_ptme.is_present':'Trouvé'}, inplace=True)

    print(f"Apel_ptme processed: {Apel_ptme.shape}")
    print(f"Visite_ptme processed: {Visite_ptme.shape}")
    print(f"Ration_ptme processed: {Ration_ptme.shape}")

    ptme = pd.concat([Apel_ptme, Visite_ptme, Ration_ptme], axis=0, ignore_index=True)
    print(f"PTME combined: {ptme.shape}")

    Apel_oev = Apel_oev[apel_oev_column]
    print(f"Apel_oev final: {Apel_oev.shape}")

    Ration_oev = Ration_oev[visite_ration_oev_column]
    print(f"Ration_oev final: {Ration_oev.shape}")

    # Combiner OEV
    Apel_oev.rename(columns={'form.appels_oev.patient_code':'patient_code',
                       'form.appels_oev.date_appel':'date',
                       'form.appels_oev.parenttuteur_trouve':'Trouvé'}, inplace=True)
    Ration_oev.rename(columns={'form.visit_ratio_and_others.patient_code':'patient_code',
                       'form.visit_ratio_and_others.date_of_visit':'date',
                       'form.visit_ratio_and_others.is_benficiary_present':'Trouvé'}, inplace=True)
    oev_visite.rename(columns={'form.visite_enfant.patient_code':'patient_code',
                       'form.visite_enfant.date_of_visit':'date',
                       'form.visite_enfant.is_available_at_time_visit':'Trouvé'}, inplace=True)

    print(f"Apel_oev renamed: {Apel_oev.shape}")
    print(f"oev_visite renamed: {oev_visite.shape}")
    print(f"Ration_oev renamed: {Ration_oev.shape}")

    oev = pd.concat([Apel_oev, oev_visite, Ration_oev], axis=0, ignore_index=True)
    print(f"OEV combined: {oev.shape}")

    data = pd.concat([ptme, oev], axis=0, ignore_index=True)
    print(f"All data combined: {data.shape}")

    data['date'].replace('---', '1901-01-01', inplace=True)

    # Pass start_date and end_date as parameters
    data = dataframe_for_period(data, 'date', start_date, end_date)
    print(f"Data after period filter: {data.shape}")

    data = transform_to_month_year_french(data, 'date')

    data.replace({'Trouvé':{0:"Non", 1:"Oui"}}, inplace=True)
    data.replace({'Trouvé':{'0':"Non", '1':"Oui"}}, inplace=True)
    print(f"Data after replacement: {data.shape}")

    data_nocorrect = data[(data['Trouvé'] == '---')]
    print(f"Data nocorrect: {data_nocorrect.shape}")

    data_cleaned = data[(data['Trouvé'] == 'Oui') | (data['Trouvé'] == 'Non')]
    print(f"Data cleaned: {data_cleaned.shape}")

    cleaned = ['formid','date','Programme','Trouvé','Type','commune', 'mois','annee','username','patient_code']
    data_cleaned = data_cleaned[cleaned]
    print(f"Data cleaned final: {data_cleaned.shape}")

    if isinstance(data, pd.DataFrame):
        grouped_data_cleaned = data_cleaned.groupby(['Programme', 'Type']).size().reset_index(name='Performances')
        print(grouped_data_cleaned)
    else:
        print("Error: 'data' is not a DataFrame.")

    # Group data for visits by monitor and calculate performances
    agent_data_visite = (
        data_cleaned[data_cleaned['Type'] == "Visite"]
        .groupby(['username'])
        .size()
        .reset_index(name='Performances')
    )

    total_visits = agent_data_visite['Performances'].sum()
    print(f"Total number of visits by monitors: {total_visits}")
    print(agent_data_visite)

    # Group data for calls by monitor and calculate performances
    agent_data_appel = (
        data_cleaned[data_cleaned['Type'] == "Appel"]
        .groupby(['username'])
        .size()
        .reset_index(name='Performances')
    )

    total_calls = agent_data_appel['Performances'].sum()
    print(f"Total number of calls by monitors: {total_calls}")
    print(agent_data_appel)

    # Group data for all performances by monitor
    agent_data = (data_cleaned.groupby(['username'])
        .size()
        .reset_index(name='Performances')
    )

    total_performances = agent_data['Performances'].sum()
    print(f"Total number of performances by monitors: {total_performances}")
    print(agent_data)

    data_pivotable = create_pivot_table(data_cleaned[data_cleaned['Type'] == "Visite"], 'formid')
    print("Data pivotable created")

    # Filter the data for Type = "Appel"
    appel_data = data_cleaned[data_cleaned['Type'] == "Appel"]

    # Create a pivot table for the filtered data
    appel_pivot_table = create_pivot_table(appel_data, 'patient_code')
    print("Appel pivot table created")

    # Filter the data for Type = "Visite"
    visit_data = data_cleaned[data_cleaned['Type'] == "Visite"]

    # Create a pivot table for the filtered data
    visit_pivot_table = create_pivot_table(visit_data, 'patient_code')
    print("Visit pivot table created")

    print('Create pivot table for total performance (Appel and Visite)')
    total_table = data_cleaned[(data_cleaned['Type'] == "Appel") | (data_cleaned['Type'] == "Visite")]
    if not total_table.empty:
        total = create_pivot_table(total_table, 'patient_code')
        if isinstance(total.columns, pd.MultiIndex) and 'Oui' in total.columns.get_level_values(1):
            total_last = add_found_percentage(total)
            print("Total performance pivot created")
        else:
            print("No 'Oui' values found in 'Trouvé' column for total performance.")
    else:
        print("No data found for total performance (Appel and Visite).")

    print('Create pivot table for OEV visit')
    visite_table_oev = data_cleaned[(data_cleaned['Programme'] == "OEV") & (data_cleaned['Type'] == "Visite")]
    if not visite_table_oev.empty:
        visite_oev = create_pivot_table(visite_table_oev, 'patient_code')
        if isinstance(visite_oev.columns, pd.MultiIndex) and 'Oui' in visite_oev.columns.get_level_values(1):
            visite_oev_last = add_found_percentage(visite_oev)
            print("OEV visit pivot created")
        else:
            print("No 'Oui' values found in 'Trouvé' column for OEV visits.")
    else:
        print("No data found for OEV visits.")

    print('Create pivot table for PTME visit')
    visite_table_ptme = data_cleaned[(data_cleaned['Programme'] == "PTME") & (data_cleaned['Type'] == "Visite")]
    if not visite_table_ptme.empty:
        visite_ptme = create_pivot_table(visite_table_ptme, 'patient_code')
        if isinstance(visite_ptme.columns, pd.MultiIndex) and 'Oui' in visite_ptme.columns.get_level_values(1):
            visite_ptme_last = add_found_percentage(visite_ptme)
            print("PTME visit pivot created")
        else:
            print("No 'Oui' values found in 'Trouvé' column for PTME visits.")
    else:
        print("No data found for PTME visits.")
     
    print('Create pivot table for call PTME')
    appel_table_ptme = data_cleaned[(data_cleaned['Programme'] == "PTME") & (data_cleaned['Type'] == "Appel")]
    if not appel_table_ptme.empty:
        appel_ptme = create_pivot_table(appel_table_ptme, 'patient_code')
        if isinstance(appel_ptme.columns, pd.MultiIndex) and 'Oui' in appel_ptme.columns.get_level_values(1):
            appel_ptme_last = add_found_percentage(appel_ptme)
            print("PTME call pivot created")
        else:
            print("No 'Oui' values found in 'Trouvé' column for PTME call.")
    else:
        print("No data found for call PTME.")
     
    print('Create pivot table for call OEV')
    appel_table_oev = data_cleaned[(data_cleaned['Programme'] == "OEV") & (data_cleaned['Type'] == "Appel")]
    if not appel_table_oev.empty:
        appel_oev = create_pivot_table(appel_table_oev, 'patient_code')
        if isinstance(appel_oev.columns, pd.MultiIndex) and 'Oui' in appel_oev.columns.get_level_values(1):
            appel_oev_last = add_found_percentage(appel_oev)
            print("OEV call pivot created")
        else:
            print("No 'Oui' values found in 'Trouvé' column for OEV call.")
    else:
        print("No data found for call OEV.")
        
        
    # Create Excel output
    print("Creating Excel output...")
    # Create a folder to store the Excel file
    output_folder = "Resultat"
    os.makedirs(output_folder, exist_ok=True)

    # ----> Ajout de la date du jour au nom de fichier
    today_str = datetime.now().strftime("%Y-%m-%d")
    output_file = os.path.join(output_folder, f"Table_des_Performances_{today_str}.xlsx")

    # Combine all pivot tables into one Excel sheet
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        sheet_name = "Tableau des Appels et Visites"  # Single sheet name

        for pivot_table, table_text, text_row in zip(
            [total_last, visite_oev_last, appel_oev_last, visite_ptme_last, appel_ptme_last],
            ["Table des Appels et Visites", "Table des visites OEV", "Table des appels OEV",
            "Table des visites PTME", "Table des appels PTME"],
            [1, 7, 13, 19, 25]
        ):
            # Texte de section
            pd.DataFrame([table_text]).to_excel(
                writer, sheet_name=sheet_name, startrow=text_row - 1, index=False, header=False
            )
            # Tableau
            pivot_table.to_excel(writer, sheet_name=sheet_name, startrow=text_row, index=True)

        # Mise en forme
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        for row, text in zip(
            [1, 7, 13, 19, 25],
            ["Table des Appels et Visites","Table des visites OEV","Table des appels OEV",
            "Table des visites PTME","Table des appels PTME"]
        ):
            worksheet.merge_cells(f"A{row}:L{row}")
            cell = worksheet[f"A{row}"]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Largeur de la première colonne (adaptez si besoin)
        worksheet.column_dimensions[get_column_letter(1)].width = 20

    print(f"All pivot tables have been saved in the '{output_file}' file.")

    # ---------- Ajout d'une feuille "Agents" dans le même fichier ----------
    def add_dataframe_to_workbook(file_path: str, sheet_name: str, df: pd.DataFrame):
        # Remplace la feuille si elle existe déjà; sinon la crée
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Écrit 'agent_data' dans le même classeur daté
    add_dataframe_to_workbook(output_file, "Agents", agent_data)

    # Personne trouvée par mois
    mois_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    # Convert the 'mois' column to a categorical type with the defined order
    data_cleaned['mois'] = pd.Categorical(data_cleaned['mois'], categories=mois_order, ordered=True)
    groupeby_data = data_cleaned.groupby(['Programme', 'Trouvé']).size().reset_index(name='Performances')
    groupby_data = groupeby_data.query('Performances > 0')
    print("Groupby data created:")
    print(groupby_data)

    # Personne trouvée par mois détaillé
    groupeby_data = data_cleaned.groupby(['mois', 'Programme', 'Trouvé']).size().reset_index(name='Performances')
    groupeby_data = groupeby_data.query('Performances > 0 and Trouvé =="Oui"')
    print("Detailed groupby data created:")
    print(groupeby_data)

    # Export individual pivot table
    export_pivot_table_to_excel(data_pivotable, 'data_pivotable.xlsx', 'Pivot Table', 'Tableau Croisé Dynamique des performances', delete_row=5)

    # Export multiple pivot tables - ORDRE SPÉCIFIQUE: "Tableau des Appels et Visites" AVANT "Listes des Appels et Visites"
    pivot_tables = [total_last, data_cleaned, groupeby_data, appel_oev_last, appel_ptme_last, visite_oev_last, visite_ptme_last, agent_data_visite, agent_data_appel]
    sheet_names = ['Tableau des Appels et Visites', 'Listes des Appels et Visites', 'Personnes trouvées', 'Appel OEV', 'Appel PTME', 'Visite OEV', 'Visite PTME', 'Visite par agent', 'Appels par agent']
    titles = ['Tableau des Appels et Visites - Résumé Principal', 'Liste Détaillée des Appels et Visites', 'Pivotable des personnes trouvées par mois/programme', 'Tableau Croisé des Appels - OEV', 'Tableau Croisé des Appels - PTME', 'Tableau Croisé des Visites - OEV', 'Tableau Croisé des Visites - PTME', 'Performances visite par agent', 'Performances Appels par agent']

    export_multiple_pivot_tables_to_excel(pivot_tables, f'Tableau_des_Performances {today_date}.xlsx', sheet_names, titles, delete_row=None)
        # Écrit 'agent_data' dans le même classeur daté
    from pathlib import Path
    from openpyxl import load_workbook, Workbook

    def copy_sheet_with_formatting(src_path: str, src_sheet: str, dst_path: str, dst_sheet: str, replace: bool = True):
        """
        Copie une feuille d'un classeur source vers un classeur destination en préservant :
        valeurs/formules, merges, largeurs de colonnes, hauteurs de lignes et styles de base.
        Si `replace=True` et que la feuille existe, elle est supprimée puis recréée.
        Positionne automatiquement le 'Tableau des Appels et Visites' en première position.
        """
        try:
            wb_src = load_workbook(src_path, data_only=False)
            if src_sheet not in wb_src.sheetnames:
                raise ValueError(f"Feuille '{src_sheet}' introuvable dans {src_path}")
            ws_src = wb_src[src_sheet]

            # S'assurer que le classeur destination existe
            dst_path_p = Path(dst_path)
            if not dst_path_p.exists():
                wb_tmp = Workbook()
                wb_tmp.save(dst_path_p)

            wb_dst = load_workbook(dst_path_p)

            # Remplacer si demandé
            if dst_sheet in wb_dst.sheetnames and replace:
                del wb_dst[dst_sheet]

            ws_dst = wb_dst.create_sheet(dst_sheet)

            # Largeurs de colonnes
            for col_letter, dim in ws_src.column_dimensions.items():
                if dim.width:
                    ws_dst.column_dimensions[col_letter].width = dim.width

            # Hauteurs de lignes
            for idx, dim in ws_src.row_dimensions.items():
                if dim.height:
                    ws_dst.row_dimensions[idx].height = dim.height

            # Geler les volets si présent
            if ws_src.freeze_panes:
                ws_dst.freeze_panes = ws_src.freeze_panes

            # Copie cellulaire (valeurs + styles) - CORRECTION PRINCIPALE
            for row in ws_src.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # Utiliser row et column au lieu de col_idx inexistant
                        tgt = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                        
                        # Copy font properties individually to avoid StyleProxy error
                        if cell.font and cell.font != ws_src['A1'].font:
                            from openpyxl.styles import Font
                            try:
                                tgt.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    vertAlign=cell.font.vertAlign,
                                    underline=cell.font.underline,
                                    strike=cell.font.strike,
                                    color=cell.font.color
                                )
                            except Exception:
                                pass  # Ignorer les erreurs de style
                        
                        # Copy alignment properties individually
                        if cell.alignment:
                            from openpyxl.styles import Alignment
                            try:
                                tgt.alignment = Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    textRotation=cell.alignment.textRotation,
                                    wrapText=cell.alignment.wrapText,
                                    shrinkToFit=cell.alignment.shrinkToFit,
                                    indent=cell.alignment.indent
                                )
                            except Exception:
                                pass
                        
                        # Copy fill properties individually
                        if cell.fill and hasattr(cell.fill, 'fill_type') and cell.fill.fill_type:
                            from openpyxl.styles import PatternFill
                            try:
                                if cell.fill.fill_type != 'none':
                                    tgt.fill = PatternFill(
                                        fill_type=cell.fill.fill_type,
                                        start_color=cell.fill.start_color,
                                        end_color=cell.fill.end_color
                                    )
                            except Exception:
                                pass
                        
                        # Copy border properties individually
                        if cell.border:
                            from openpyxl.styles import Border
                            try:
                                tgt.border = Border(
                                    left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom,
                                    diagonal=cell.border.diagonal,
                                    diagonal_direction=cell.border.diagonal_direction,
                                    vertical=cell.border.vertical,
                                    horizontal=cell.border.horizontal
                                )
                            except Exception:
                                pass
                        
                        # Copy other properties
                        try:
                            if cell.number_format and cell.number_format != 'General':
                                tgt.number_format = cell.number_format
                        except Exception:
                            pass
                        
                        # Copy protection properties individually
                        if cell.protection:
                            from openpyxl.styles import Protection
                            try:
                                tgt.protection = Protection(
                                    locked=cell.protection.locked,
                                    hidden=cell.protection.hidden
                                )
                            except Exception:
                                pass

            # Fusions
            try:
                for m in ws_src.merged_cells.ranges:
                    ws_dst.merge_cells(str(m))
            except Exception as e:
                print(f"Erreur lors de la copie des cellules fusionnées: {e}")

            # Mise en page (optionnel)
            try:
                ws_dst.page_setup = ws_src.page_setup
                ws_dst.page_margins = ws_src.page_margins
                ws_dst.print_options = ws_src.print_options
            except Exception:
                pass  # Ignorer les erreurs de mise en page

            # AUTOMATIQUEMENT PLACER "Tableau des Appels et Visites" EN PREMIÈRE POSITION
            # Assurer que cette feuille est toujours en première position et avant "Listes des Appels et Visites"
            all_sheets = list(wb_dst.sheetnames)
            if dst_sheet in all_sheets:
                # Si le nom de la feuille contient "Tableau des Appels et Visites"
                if "Tableau des Appels et Visites" in dst_sheet:
                    # Si ce n'est pas déjà la première feuille
                    if all_sheets[0] != dst_sheet:
                        target_sheet = wb_dst[dst_sheet]
                        # Déplacer en première position
                        wb_dst.move_sheet(target_sheet, offset=-len(all_sheets))
                        print(f"✅ Feuille '{dst_sheet}' déplacée en première position.")
                    else:
                        print(f"✅ Feuille '{dst_sheet}' déjà en première position.")
                    
                    # Vérifier que "Listes des Appels et Visites" est après si elle existe
                    if "Listes des Appels et Visites" in all_sheets:
                        updated_sheets = list(wb_dst.sheetnames)
                        tableau_index = updated_sheets.index(dst_sheet)
                        liste_index = updated_sheets.index("Listes des Appels et Visites")
                        
                        if liste_index < tableau_index:
                            # Déplacer "Listes des Appels et Visites" après "Tableau des Appels et Visites"
                            liste_sheet = wb_dst["Listes des Appels et Visites"]
                            wb_dst.move_sheet(liste_sheet, offset=(tableau_index - liste_index))
                            print(f"✅ Feuille 'Listes des Appels et Visites' déplacée après '{dst_sheet}'.")
                        else:
                            print(f"✅ L'ordre correct est respecté: '{dst_sheet}' avant 'Listes des Appels et Visites'.")

            wb_dst.save(dst_path_p)
            wb_src.close()
            wb_dst.close()
            
        except Exception as e:
            print(f"Erreur dans copy_sheet_with_formatting: {e}")
            raise

    # -------------------------
    # Paramétrage demandé
    # -------------------------
    today_str = datetime.now().strftime("%Y-%m-%d")

    src_path = os.path.join("Resultat", f"Table_des_Performances_{today_str}.xlsx")
    src_sheet = "Tableau des Appels et Visites"

    dst_path = f"./Tableau_des_Performances {today_date}.xlsx"
    dst_sheet = "Tableau des Appels et Visites"

    # Exécution
    copy_sheet_with_formatting(
        src_path=src_path,
        src_sheet=src_sheet,
        dst_path=dst_path,
        dst_sheet=dst_sheet,
        replace=True
    )
    print(f"Feuille '{src_sheet}' copiée vers '{dst_path}' en tant que '{dst_sheet}'.")
    
    print(f"Tableau des performances créé avec 'Tableau des Appels et Visites' en première position.")

    # Additional data processing
    nut_column = ['form.date_de_visite','username','Programme','Type']
    eva_column = ['form.Evaluation_agent.date_devaluation_fra','form.Evaluation_agent.nom_du_superviseur','Programme','Type']

    print("CallApp Analysis completed successfully!")

if __name__ == "__main__":
    main()